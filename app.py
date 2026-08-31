"""
Article Generator
------------------
A simple desktop app that:
  1. Reads a list of topics from an Excel file you upload
  2. Asks Claude to write a full article for each topic
  3. Asks Claude to suggest images for the article
  4. Generates those images using OpenAI's image API
  5. Saves everything (a .blade.php file + images) into a folder
     named after the topic, inside a folder you choose (Desktop by default)

You do not need to know how to code to use this. Just run the app,
paste in your API keys once, choose your Excel file, and click Start.
"""

import json
import os
import re
import sys
import base64
import traceback
import threading
import queue
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import requests
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Basic settings
# ---------------------------------------------------------------------------

APP_DIR = Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "config.json"

ANTHROPIC_MODEL = "claude-sonnet-5"
ANTHROPIC_API_URL = "https://api.anthropic.com/v1/messages"
ANTHROPIC_VERSION = "2023-06-01"

OPENAI_IMAGE_MODEL = "gpt-image-1"
OPENAI_IMAGE_URL = "https://api.openai.com/v1/images/generations"

DEFAULT_OUTPUT_DIR = str(Path.home() / "Desktop")


# ---------------------------------------------------------------------------
# Config (API keys) — saved locally so you only type them in once
# ---------------------------------------------------------------------------

def load_config():
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text())
        except Exception:
            return {}
    return {}


def save_config(data):
    CONFIG_PATH.write_text(json.dumps(data, indent=2))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def slugify(text):
    """Turn 'How to Brew Coffee?' into 'how-to-brew-coffee'."""
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9\s-]", "", text)
    text = re.sub(r"[\s_]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "untitled"


def unique_folder(base_dir: Path, slug: str) -> Path:
    """Avoid overwriting: if 'my-topic' exists, use 'my-topic-2', etc."""
    candidate = base_dir / slug
    counter = 2
    while candidate.exists():
        candidate = base_dir / f"{slug}-{counter}"
        counter += 1
    return candidate


def read_excel_topics(path):
    """
    Reads the uploaded Excel file. Expects a header row with a column
    called 'Topic' (case-insensitive). Optional columns: 'Keywords',
    'Notes' / 'Outline'. Any other columns are ignored.
    Returns a list of dicts: [{"topic": ..., "keywords": ..., "notes": ...}, ...]
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The Excel file appears to be empty.")

    header = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col_index(*names):
        for name in names:
            if name in header:
                return header.index(name)
        return None

    topic_idx = col_index("topic", "title")
    keywords_idx = col_index("keywords", "keyword")
    notes_idx = col_index("notes", "outline", "description")

    if topic_idx is None:
        raise ValueError(
            "Couldn't find a 'Topic' column in the Excel file. "
            "Please make sure the first row has a header called 'Topic'."
        )

    topics = []
    for row in rows[1:]:
        if row is None or topic_idx >= len(row):
            continue
        topic_val = row[topic_idx]
        if not topic_val or not str(topic_val).strip():
            continue
        topics.append({
            "topic": str(topic_val).strip(),
            "keywords": str(row[keywords_idx]).strip()
                if keywords_idx is not None and keywords_idx < len(row) and row[keywords_idx] else "",
            "notes": str(row[notes_idx]).strip()
                if notes_idx is not None and notes_idx < len(row) and row[notes_idx] else "",
        })
    return topics


# ---------------------------------------------------------------------------
# Claude API calls
# ---------------------------------------------------------------------------

def call_claude(api_key, prompt, max_tokens=4000):
    headers = {
        "x-api-key": api_key,
        "anthropic-version": ANTHROPIC_VERSION,
        "content-type": "application/json",
    }
    body = {
        "model": ANTHROPIC_MODEL,
        "max_tokens": max_tokens,
        "messages": [{"role": "user", "content": prompt}],
    }
    resp = requests.post(ANTHROPIC_API_URL, headers=headers, json=body, timeout=120)
    if resp.status_code != 200:
        raise RuntimeError(f"Claude API error ({resp.status_code}): {resp.text[:500]}")
    data = resp.json()
    text_parts = [b["text"] for b in data.get("content", []) if b.get("type") == "text"]
    return "".join(text_parts).strip()


def strip_json_fences(text):
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*\n?", "", text)
        text = re.sub(r"```$", "", text.strip())
    return text.strip()


def generate_article(api_key, topic, keywords, notes, num_images, article_context=""):
    context_block = ""
    if article_context.strip():
        context_block = f"""
Reference context (pasted from a previous conversation — use it for style,
tone, facts, or background info where relevant; ignore anything irrelevant
to this specific topic):
---
{article_context.strip()}
---
"""

    prompt = f"""You are an expert content writer. Write a complete, well-structured
article based on the details below.

Topic: {topic}
Keywords to naturally include: {keywords or "(none given)"}
Additional notes / outline to follow: {notes or "(none given)"}
{context_block}
Requirements:
- 700-1100 words
- Use an engaging title
- Use HTML formatting for the body: <h2> for section headings, <p> for
  paragraphs, <ul>/<li> for lists where useful. Do not include <html>,
  <head>, or <body> tags — just the inner content.
- Also write a one-sentence meta description (under 160 characters) for SEO.
- Suggest exactly {num_images} images that should accompany this article.
  For each image, give a short filename-friendly label (2-4 words, lowercase,
  hyphenated) and a detailed visual description suitable for an AI image
  generator.

Respond with ONLY a JSON object (no markdown fences, no extra text) in
exactly this shape:
{{
  "title": "...",
  "meta_description": "...",
  "content_html": "...",
  "images": [
    {{"label": "...", "prompt": "..."}}
  ]
}}
"""
    raw = call_claude(api_key, prompt, max_tokens=4000)
    raw = strip_json_fences(raw)
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"Claude didn't return valid JSON for '{topic}': {e}\nRaw reply:\n{raw[:800]}")


# ---------------------------------------------------------------------------
# OpenAI image generation
# ---------------------------------------------------------------------------

def generate_image(api_key, prompt, save_path, image_context=""):
    if image_context.strip():
        prompt = (
            f"{prompt}\n\n"
            f"Style/reference notes (from a previous conversation, apply where relevant): "
            f"{image_context.strip()}"
        )
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    body = {
        "model": OPENAI_IMAGE_MODEL,
        "prompt": prompt,
        "size": "1024x1024",
        "n": 1,
    }
    resp = requests.post(OPENAI_IMAGE_URL, headers=headers, json=body, timeout=120)
    if resp.status_code != 200:
        raise RuntimeError(f"Image API error ({resp.status_code}): {resp.text[:500]}")
    data = resp.json()
    item = data["data"][0]

    if "b64_json" in item and item["b64_json"]:
        img_bytes = base64.b64decode(item["b64_json"])
    elif "url" in item and item["url"]:
        img_resp = requests.get(item["url"], timeout=120)
        img_resp.raise_for_status()
        img_bytes = img_resp.content
    else:
        raise RuntimeError("Image API returned no image data.")

    save_path.write_bytes(img_bytes)


# ---------------------------------------------------------------------------
# Blade file builder
# ---------------------------------------------------------------------------

def build_blade_file(title, meta_description, content_html, image_filenames):
    images_block = ""
    for fname in image_filenames:
        images_block += f'    <img src="{{{{ asset(\'images/{fname}\') }}}}" alt="{title}">\n'

    return f"""{{{{--
    Title: {title}
    Meta description: {meta_description}
    Auto-generated by Article Generator
--}}}}

<article>
    <h1>{title}</h1>

{images_block}
{content_html}
</article>
"""


# ---------------------------------------------------------------------------
# Core pipeline for a single row
# ---------------------------------------------------------------------------

def process_topic(row, output_root: Path, anthropic_key, openai_key, num_images, log,
                   article_context="", image_context=""):
    topic = row["topic"]
    log(f"Writing article for: {topic}")

    article = generate_article(anthropic_key, topic, row["keywords"], row["notes"], num_images, article_context)

    slug = slugify(article.get("title") or topic)
    folder = unique_folder(output_root, slug)
    folder.mkdir(parents=True, exist_ok=True)

    images = article.get("images", [])[:num_images]
    image_filenames = []

    if openai_key:
        for i, img in enumerate(images, start=1):
            img_slug = slugify(img.get("label", f"image-{i}"))
            fname = f"{img_slug}.png"
            log(f"  Generating image {i}/{len(images)}: {img.get('label', '')}")
            try:
                generate_image(openai_key, img.get("prompt", topic), folder / fname, image_context)
                image_filenames.append(fname)
            except Exception as e:
                log(f"  Image {i} failed, skipping: {e}")
    else:
        # No image key provided — just save the suggestions as text
        suggestions_text = "\n\n".join(
            f"{img.get('label', 'image')}: {img.get('prompt', '')}" for img in images
        )
        (folder / "image-suggestions.txt").write_text(suggestions_text)
        log("  No OpenAI key set — saved image suggestions as text instead.")

    blade_content = build_blade_file(
        article.get("title", topic),
        article.get("meta_description", ""),
        article.get("content_html", ""),
        image_filenames,
    )
    blade_path = folder / f"{slug}.blade.php"
    blade_path.write_text(blade_content, encoding="utf-8")

    log(f"  Done -> {folder}")
    return folder


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Article Generator")
        self.geometry("700x760")
        self.resizable(True, True)

        self.config_data = load_config()
        self.excel_path = tk.StringVar()
        self.output_dir = tk.StringVar(value=self.config_data.get("output_dir", DEFAULT_OUTPUT_DIR))
        self.anthropic_key = tk.StringVar(value=self.config_data.get("anthropic_key", ""))
        self.openai_key = tk.StringVar(value=self.config_data.get("openai_key", ""))
        self.num_images = tk.StringVar(value=str(self.config_data.get("num_images", 2)))
        self._saved_article_context = self.config_data.get("article_context", "")
        self._saved_image_context = self.config_data.get("image_context", "")

        self.stop_flag = threading.Event()
        self.worker_thread = None
        self.msg_queue = queue.Queue()

        self._build_ui()
        self.after(150, self._poll_queue)

    # -- UI layout -----------------------------------------------------

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        settings = ttk.LabelFrame(self, text="1. Settings (one-time setup)")
        settings.pack(fill="x", **pad)

        ttk.Label(settings, text="Anthropic (Claude) API key:").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        ttk.Entry(settings, textvariable=self.anthropic_key, show="*", width=50).grid(row=0, column=1, padx=8, pady=4)

        ttk.Label(settings, text="OpenAI API key (optional, for images):").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        ttk.Entry(settings, textvariable=self.openai_key, show="*", width=50).grid(row=1, column=1, padx=8, pady=4)

        ttk.Label(settings, text="Images per article:").grid(row=2, column=0, sticky="w", padx=8, pady=4)
        ttk.Combobox(settings, textvariable=self.num_images, values=["1", "2", "3"], width=5, state="readonly").grid(row=2, column=1, sticky="w", padx=8, pady=4)

        ttk.Button(settings, text="Save Settings", command=self._save_settings).grid(row=3, column=0, columnspan=2, pady=8)

        # -- Reference context (pasted from existing chats) --
        context = ttk.LabelFrame(self, text="2. Reference context (optional) — paste from an existing Claude/ChatGPT chat")
        context.pack(fill="x", **pad)

        ttk.Label(
            context,
            text="Copy text from a Claude chat and paste it below to guide the article's\n"
                 "style, tone, or background facts:",
            justify="left",
        ).grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
        self.article_context_text = tk.Text(context, height=4, width=70, wrap="word")
        self.article_context_text.grid(row=1, column=0, padx=8, pady=(0, 8))
        self.article_context_text.insert("1.0", self._saved_article_context)

        ttk.Label(
            context,
            text="Copy text/description from a ChatGPT chat and paste it below to guide\n"
                 "the image style (colors, mood, art style, etc.):",
            justify="left",
        ).grid(row=2, column=0, sticky="w", padx=8, pady=(0, 2))
        self.image_context_text = tk.Text(context, height=4, width=70, wrap="word")
        self.image_context_text.grid(row=3, column=0, padx=8, pady=(0, 8))
        self.image_context_text.insert("1.0", self._saved_image_context)

        # -- File selection --
        files = ttk.LabelFrame(self, text="3. Choose files")
        files.pack(fill="x", **pad)

        ttk.Button(files, text="Choose Excel File...", command=self._choose_excel).grid(row=0, column=0, padx=8, pady=6)
        ttk.Label(files, textvariable=self.excel_path, wraplength=450).grid(row=0, column=1, sticky="w", padx=8)

        ttk.Button(files, text="Choose Output Folder...", command=self._choose_output).grid(row=1, column=0, padx=8, pady=6)
        ttk.Label(files, textvariable=self.output_dir, wraplength=450).grid(row=1, column=1, sticky="w", padx=8)

        # -- Run controls --
        run_frame = ttk.LabelFrame(self, text="4. Run")
        run_frame.pack(fill="x", **pad)

        btn_row = ttk.Frame(run_frame)
        btn_row.pack(fill="x", padx=8, pady=6)
        self.start_btn = ttk.Button(btn_row, text="Start", command=self._start)
        self.start_btn.pack(side="left")
        self.stop_btn = ttk.Button(btn_row, text="Stop", command=self._stop, state="disabled")
        self.stop_btn.pack(side="left", padx=8)

        self.progress = ttk.Progressbar(run_frame, mode="determinate")
        self.progress.pack(fill="x", padx=8, pady=6)

        self.status_label = ttk.Label(run_frame, text="Ready.")
        self.status_label.pack(anchor="w", padx=8)

        # -- Log --
        log_frame = ttk.LabelFrame(self, text="Log")
        log_frame.pack(fill="both", expand=True, **pad)

        self.log_text = tk.Text(log_frame, height=8, state="disabled", wrap="word")
        self.log_text.pack(fill="both", expand=True, padx=8, pady=8)

    # -- Actions ---------------------------------------------------------

    def _save_settings(self):
        self.config_data.update({
            "anthropic_key": self.anthropic_key.get().strip(),
            "openai_key": self.openai_key.get().strip(),
            "num_images": int(self.num_images.get()),
            "output_dir": self.output_dir.get().strip(),
            "article_context": self.article_context_text.get("1.0", "end").strip(),
            "image_context": self.image_context_text.get("1.0", "end").strip(),
        })
        save_config(self.config_data)
        messagebox.showinfo("Saved", "Settings saved. You won't need to re-enter them next time.")

    def _choose_excel(self):
        path = filedialog.askopenfilename(
            title="Choose your Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm")],
        )
        if path:
            self.excel_path.set(path)

    def _choose_output(self):
        path = filedialog.askdirectory(title="Choose where to save the results", initialdir=self.output_dir.get())
        if path:
            self.output_dir.set(path)

    def _log(self, message):
        self.msg_queue.put(("log", message))

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.msg_queue.get_nowait()
                if kind == "log":
                    self.log_text.configure(state="normal")
                    self.log_text.insert("end", payload + "\n")
                    self.log_text.see("end")
                    self.log_text.configure(state="disabled")
                elif kind == "progress":
                    self.progress["value"] = payload
                elif kind == "status":
                    self.status_label.configure(text=payload)
                elif kind == "done":
                    self.start_btn.configure(state="normal")
                    self.stop_btn.configure(state="disabled")
        except queue.Empty:
            pass
        self.after(150, self._poll_queue)

    def _start(self):
        if not self.excel_path.get():
            messagebox.showwarning("Missing file", "Please choose an Excel file first.")
            return
        if not self.anthropic_key.get().strip():
            messagebox.showwarning("Missing API key", "Please enter your Anthropic (Claude) API key and click Save Settings.")
            return

        self._save_settings()
        self.stop_flag.clear()
        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.progress["value"] = 0

        self.worker_thread = threading.Thread(target=self._run_pipeline, daemon=True)
        self.worker_thread.start()

    def _stop(self):
        self.stop_flag.set()
        self._log("Stopping after the current article finishes...")

    def _run_pipeline(self):
        try:
            topics = read_excel_topics(self.excel_path.get())
        except Exception as e:
            self._log(f"Error reading Excel file: {e}")
            self.msg_queue.put(("done", None))
            return

        if not topics:
            self._log("No topics found in the Excel file.")
            self.msg_queue.put(("done", None))
            return

        total = len(topics)
        self._log(f"Found {total} topic(s). Starting...")
        output_root = Path(self.output_dir.get())
        output_root.mkdir(parents=True, exist_ok=True)

        anthropic_key = self.anthropic_key.get().strip()
        openai_key = self.openai_key.get().strip()
        num_images = int(self.num_images.get())
        article_context = self.article_context_text.get("1.0", "end").strip()
        image_context = self.image_context_text.get("1.0", "end").strip()

        succeeded, failed = 0, 0

        for i, row in enumerate(topics, start=1):
            if self.stop_flag.is_set():
                self._log("Stopped by user.")
                break

            self.msg_queue.put(("status", f"Processing {i} of {total}: {row['topic']}"))
            try:
                process_topic(row, output_root, anthropic_key, openai_key, num_images, self._log,
                               article_context, image_context)
                succeeded += 1
            except Exception as e:
                failed += 1
                self._log(f"FAILED on '{row['topic']}': {e}")
                self._log(traceback.format_exc(limit=1))

            self.msg_queue.put(("progress", int(i / total * 100)))

        self.msg_queue.put(("status", f"Finished. {succeeded} succeeded, {failed} failed."))
        self._log(f"All done. {succeeded} succeeded, {failed} failed.")
        self.msg_queue.put(("done", None))


if __name__ == "__main__":
    app = App()
    app.mainloop()
